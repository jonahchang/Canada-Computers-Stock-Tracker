import time
import os
import sys
import json
import tkinter as tk
from tkinter import Tk, filedialog, simpledialog, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DEBUG_MODE = True

json_file = "CC_products.json"
original_file = "CC_original_products.json"

def load_products():
    global power_supplies, coolers, chassis, external

    # Load product list
    if os.path.exists(json_file):
        with open(json_file, "r") as f:
            data = json.load(f)
            power_supplies = data.get("power_supplies", {})
            coolers = data.get("coolers", {})
            chassis = data.get("chassis", {})
            external = data.get("miscellaneous", {})
    else:
        # Chassis
        chassis = {"GT301/BLK/ARGB FAN": "https://www.canadacomputers.com/en/mid-tower-cases/168147/asus-tuf-gaming-gt301-mid-tower-compact-case-for-atx-motherboards-gt301-blk-argb-fan.html",
                    "GT502/WHT/TG//": "https://www.canadacomputers.com/en/mid-tower-cases/233894/asus-tuf-gaming-gt502-white-atx-mid-tower-computer-case-gt502-wht-tg-.html",
                    "GT502/BLK/TG//": "https://www.canadacomputers.com/en/mid-tower-cases/233895/asus-tuf-gaming-gt502-atx-mid-tower-computer-case-with-front-panel-rgb-button-gt502-blk-tg-.html",
                    "GR101 ROG Z11 CASE/BLK": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/240012/asus-rog-gr101-z11-mini-itx-dtx-tempered-glass-gaming-case-gr101-rog-z11-case-blk.html",
                    "AP201 BLK MESH": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/243435/asus-prime-ap201-microatx-mesh-small-tower-case-black-ap201-blk-mesh.html", 
                    "AP201 WHT MESH": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/243436/asus-prime-ap201-microatx-mesh-small-tower-case-white-ap201-wht-mesh.html",
                    "AP201 ASUS PRIME CASE TG BLACK": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/244029/asus-prime-ap201-microatx-tempered-glass-small-tower-case-black-ap201-asus-prime-case-tg-black.html",
                    "AP201 WHT TG": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/244030/asus-prime-ap201-microatx-tempered-glass-small-tower-case-white-ap201-wht-tg.html",
                    "GR701/BK/PWM FAN//": "https://www.canadacomputers.com/en/full-tower-cases/244527/asus-rog-hyperion-gr701-e-atx-full-tower-computer-case-gr701-bk-pwm-fan-.html",
                    "GR701/WT/PWM FAN//": "https://www.canadacomputers.com/en/full-tower-cases/246209/asus-rog-hyperion-gr701-e-atx-full-tower-computer-case-white-gr701-wt-pwm-fan-.html",
                    "A21/BLK//": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/250263/asus-a21-micro-atx-case-black-a21-blk-.html",
                    "A21/WHT//": "https://www.canadacomputers.com/en/small-form-factor-mini-itx-cases/250264/asus-a21-micro-atx-case-white-a21-wht-.html",
                    "PA602 ProArt Case": "https://www.canadacomputers.com/en/mid-tower-cases/252356/asus-proart-pa602-e-atx-computer-case-pa602-proart-case.html",
                    "GT302/ARGB FANS/BLK//": "https://www.canadacomputers.com/en/mid-tower-cases/255067/asus-tuf-gaming-gt302-argb-atx-mid-tower-case-black-gt302-argb-fans-blk-.html",
                    "GT302/ARGB FANS/WHT//": "https://www.canadacomputers.com/en/mid-tower-cases/255068/asus-tuf-gaming-gt302-argb-atx-mid-tower-case-white-gt302-argb-fans-wht-.html",
                    "PA401/BK/WOOD/TG//": "https://www.canadacomputers.com/en/mid-tower-cases/268879/asus-proart-pa401-wood-edition-atx-computer-case-pa401-bk-wood-tg-.html",
                    "A31 TG Black": "https://www.canadacomputers.com/en/mid-tower-cases/269020/asus-a31-atx-mid-tower-gaming-case-dual-sided-tempered-glass-black-a31-tg-black.html"
                    }
        
        # Coolers
        coolers = {"ROG RYUO III 240 ARGB": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/237226/asus-rog-ryuo-iii-240-argb-all-in-one-liquid-cpu-cooler-rog-ryuo-iii-240-argb.html",
                    "ROG RYUO III 240 ARGB WHT": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/237231/asus-rog-ryuo-iii-240-argb-all-in-one-liquid-cpu-cooler-white-rog-ryuo-iii-240-argb-wht.html",
                    "ROG RYUO III 360 ARGB WHT": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/237235/asus-rog-ryuo-iii-360-argb-all-in-one-liquid-cpu-cooler-white-rog-ryuo-iii-360-argb-wht.html",
                    "ROG RYUJIN III 360 ARGB WHT": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/243316/asus-rog-ryujin-iii-360-argb-aio-liquid-cpu-cooler-rog-ryujin-iii-360-argb.html",
                    "ROG RYUJIN III 240 ARGB": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/243418/asus-rog-ryujin-iii-240-argb-aio-liquid-cpu-cooler-rog-ryujin-iii-240-argb.html",
                    "ROG RYUJIN III 240 ARGB WHT": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/243419/asus-rog-ryujin-iii-240-argb-aio-liquid-cpu-cooler-white-rog-ryujin-iii-240-argb-wht.html",
                    "ROG RG-07 KIT": "https://www.canadacomputers.com/en/thermal-compound/252353/asus-rog-rg-07-performance-thermal-paste-kit-3-grams-rog-rg-07-kit.html",
                    "ROG RYUJIN III 240": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/253683/asus-rog-ryujin-iii-240-all-in-one-liquid-cpu-cooler-black-rog-ryujin-iii-240.html",
                    "ROG RYUJIN III 360": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/253688/asus-rog-ryujin-iii-360-all-in-one-liquid-cpu-cooler-black-rog-ryujin-iii-360.html",
                    "ROG STRIX LC III 240 ARGB": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/254420/asus-rog-strix-lc-iii-240-argb-all-in-one-liquid-cpu-cooler-rog-strix-lc-iii-240-argb.html",
                    "ROG RYUJIN III 360 ARGB EXTREME": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/269826/asus-rog-ryujin-iii-360-argb-extreme-all-in-one-aio-cpu-liquid-cooler-rog-ryujin-iii-360-argb-extreme.html",
                    "ROG RYUJIN III 360 ARGB EXTREME WHT": "https://www.canadacomputers.com/en/aio-cpu-liquid-coolers/269827/asus-rog-ryujin-iii-360-argb-extreme-white-edition-all-in-one-aio-cpu-liquid-cooler-rog-ryujin-iii-360-argb-extreme-wht.html"
                    }
        
        # External
        external = {"ESD-S1C/BLK/G/AS": "https://www.canadacomputers.com/en/hard-drive-ssd-enclosures/165368/asus-rog-strix-arion-drive-enclosure-esd-s1c-blk-g-as.html",
                    "DRW-24B1ST/BLK/B/AS": "https://www.canadacomputers.com/en/dvd-writers/60114/asus-drw-24f1st-blk-b-internal-24x-dvd-writer-oem-black-drw-24f1st-blk-b.html",
                    "SDRW-08D2S-U/BLK/G/AS": "https://www.canadacomputers.com/en/external-blu-ray-cd-dvd-drives/36390/asus-slim-external-8x-dvd-writer-usb-2-0-black-sdrw-08d2s-u-sdrw-08d2s-u-blk-g-as.html",
                    "SBC-06D2X-U/BLK/G/AS": 0,
                    "SBW-06D2X-U/BLK/G/AS": "https://www.canadacomputers.com/en/external-blu-ray-cd-dvd-drives/47109/asus-sbw-06d2x-u-external-slim-6x-blu-ray-writer-retail-sbw-06d2x-u-blk-g-as.html",
                    "DRW-24F1ST/BLK/B": "https://www.canadacomputers.com/en/dvd-writers/60114/asus-drw-24f1st-blk-b-internal-24x-dvd-writer-oem-black-drw-24f1st-blk-b.html",
                    "BW-16D1HT": "https://www.canadacomputers.com/en/blu-ray-writers/62771/asus-bw-16d1ht-internal-16x-bdxl-blu-ray-writer-retail-box-black-bw-16d1ht.html",
                    "SDRW-08D2S-U/WHT/G/AS": "https://www.canadacomputers.com/en/external-blu-ray-cd-dvd-drives/102067/asus-sdrw-08d2s-u-slim-external-8x-dvd-writer-retail-sdrw-08d2s-u-wht-g-as.html",
                    "SBC-06D2X-U/BLK/G/AS/CA": 0,
                    "BW-16D1X-U": "https://www.canadacomputers.com/en/external-blu-ray-cd-dvd-drives/138835/asus-bw-16d1x-u-external-16x-blu-ray-writer-usb-3-0-black-bw-16d1x-u.html",
                    "DRW-24B3ST": "https://www.canadacomputers.com/en/dvd-writers/265197/asus-drw-24b3st-internal-24x-dvd-r-writing-speed-black-drw-24b3st.html"
                    }

        # Power Supplies
        power_supplies = {"ROG-STRIX-1000G": "https://www.canadacomputers.com/en/power-supplies/188790/asus-rog-strix-1000w-gold-psu-power-supply-rog-heatsinks-axial-tech-fan-design-dual-ball-fan-bearings-0db-technology-80-plus-go-rog-strix-1000g.html",
                        "ROG-THOR-1000P2-GAMING": "https://www.canadacomputers.com/en/power-supplies/215084/asus-rog-thor-rog-thor-1000p2-gaming-1000-w-atx12v-80-plus-platinum-certified-power-supply-rog-thor-1000p2-gaming.html",
                        "ROG-THOR-1600T-GAMING": "https://www.canadacomputers.com/en/power-supplies/235083/asus-rog-thor-1600w-titanium-1600-watt-fully-modular-power-supply-80-titanium-lambda-a-certified-gan-mosfets-digital-power-cont-rog-thor-1600t-gaming.html",
                        "TUF-GAMING-850G": "https://www.canadacomputers.com/en/power-supplies/239504/asus-tuf-gaming-850w-gold-power-supply-tuf-gaming-850g.html",
                        "TUF-GAMING-1000G": "https://www.canadacomputers.com/en/power-supplies/239506/asus-tuf-gaming-1000w-gold-power-supply-tuf-gaming-1000g.html",
                        "TUF-GAMING-750G": "https://www.canadacomputers.com/en/power-supplies/239507/asus-tuf-gaming-750w-gold-power-supply-tuf-gaming-750g.html",
                        "ROG-THOR-850P2-GAMING": "https://www.canadacomputers.com/en/power-supplies/239524/asus-rog-thor-850w-platinum-ii-rog-thor-850p2-gaming.html",
                        "ROG-LOKI-850P-SFX-L-GAMING": "https://www.canadacomputers.com/en/power-supplies/243485/asus-rog-loki-series-sfx-l-850w-power-supply-80-platinum-atx-3-0-compatible-rog-loki-850p-sfx-l-gaming.html",
                        "ROG-STRIX-850G": "https://www.canadacomputers.com/en/power-supplies/249138/the-rog-strix-850w-80-gold-rog-strix-850g.html",
                        "ROG-STRIX-850G-WHITE": "https://www.canadacomputers.com/en/power-supplies/249140/the-rog-strix-850w-80-gold-rog-strix-850g-white.html",
                        "AP-850G": "https://www.canadacomputers.com/en/power-supplies/253007/asus-prime-850w-gold-power-supply-ap-850g.html",
                        "AP-750G": "https://www.canadacomputers.com/en/power-supplies/253022/asus-prime-750w-gold-power-supply-ap-750g.html",
                        "ROG-STRIX-1000G-AURA-GAMING": "https://www.canadacomputers.com/en/power-supplies/267287/asus-rog-strix-1000w-gold-aura-edition-fully-modular-power-supply-80-gold-certified-atx-3-0-rog-strix-1000g-aura-gaming.html",
                        "ROG-STRIX-850G-AURA-GAMING": "https://www.canadacomputers.com/en/power-supplies/266676/asus-rog-strix-850w-gold-aura-edition-fully-modular-power-supply-80-gold-certified-atx-3-0-rog-strix-850g-aura-gaming.html",
                        
                        # wrong SKU
                        "ROG-STRIX-750G-AURA-GAMING": "https://www.canadacomputers.com/en/power-supplies/253675/asus-rog-strix-750w-gold-aura-edition-fully-modular-power-supply-80-gold-certified-atx-3-0-90ye00p3-bvaa00.html?srsltid=AfmBOopk52uR4Lnwwl6fGTNT4XklYuUb5S8Ur5WjtiR7xq8JBEcd3Mlx",
                        "TUF-GAMING-1200G": "https://www.canadacomputers.com/en/power-supplies/265984/asus-tuf-gaming-1200w-gold-power-supply-tuf-gaming-1200g.html",
                        "ROG-STRIX-1200P-GAMING": "https://www.canadacomputers.com/en/power-supplies/268226/asus-rog-strix-1200w-platinum-fully-modular-power-supply-80-platinum-certified-atx-3-1-rog-strix-1200p-gaming.html",
                        "ROG-STRIX-1000P-GAMING": "https://www.canadacomputers.com/en/power-supplies/268228/asus-rog-strix-1000w-platinum-fully-modular-power-supply-80-platinum-certified-atx-3-1-rog-strix-1000p-gaming.html"
                        }

        save_products(original_file)

        save_products(json_file)

def save_products(filename):
    with open(filename, "w") as f:
        json.dump({
            "power_supplies": power_supplies,
            "coolers": coolers,
            "chassis": chassis,
            "external": external
        }, f, indent=2)

# Mapping of store ID to store name (must match Excel header format)
store_map = [
    "Surrey", "Richmond", "Burnaby", "Coquitlam", "Vancouver Broadway", "London Masonville", "Waterloo", "Cambridge", "Hamilton", 
    "Burlington", "Barrie", "Brampton", "Oakville",	"Mississauga", "Etobicoke", "Vaughan", "Newmarket", "Richmond Hill", "North York",
    "Toronto Down Town 284", "Markham Unionville", "Toronto Kennedy", "St Catharines", "Ajax", "Whitby", "Oshawa", "Kingston", "Kanata",
    "Ottawa Merivale", "Ottawa Downtown", "Ottawa Orleans", "Gatineau", "West Island", "Laval", "Marche Central", "Montreal", "Brossard",
    "QC Vanier", "Halifax"
]

# Stock highlighting colors
green_fill = PatternFill(start_color='00FF00', end_color='00FF00',  fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00',  fill_type='solid')
blue_fill = PatternFill(start_color='83CCEB', end_color='83CCEB',  fill_type='solid')

def process_add_products_sheet(wb):
    if "Add Products" not in wb.sheetnames:
        return
    
    sheet = wb["Add Products"]
    row = 3
    added_any = False

    while True:
        name_cell = sheet[f"A{row}"]
        url_cell = sheet[f"B{row}"]
        cat_cell = sheet[f"C{row}"]

        name = name_cell.value.strip() if name_cell.value else ""
        url = url_cell.value.strip() if url_cell.value else ""
        category = (cat_cell.value or "").strip().lower()

        if not name and not url and not category:
            break

        if not name or not url:
            row += 1
            continue

        if not category:
            category = "miscellaneous"

        category_map = {
            "chassis": chassis,
            "cooler": coolers,
            "external": external,
            "power supply": power_supplies
        }

        target_dict = category_map.get(category.lower(), external)
        if name not in target_dict:
            target_dict[name] = url
            added_any = True

        for col in ["A", "B", "C"]:
            sheet[f"{col}{row}"].value = None

        row += 1

    rows = list(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3))
    clean_data = [[cell.value for cell in row] for row in rows if any(cell.value for cell in row)]

    for i in range(3, sheet.max_row + 1):
        for col in ["A", "B", "C"]:
            sheet[f"{col}{i}"].value = None

    for idx, (name, url, category) in enumerate(clean_data, start=3):
        sheet[f"A{idx}"] = name
        sheet[f"B{idx}"] = url
        sheet[f"C{idx}"] = category

    if added_any:
        save_products(json_file)

def normalize_cell_value(value):
    if value is None:
        return 0
    val = str(value).strip().upper()
    if val == 'NO STOCK':
        return 0
    if val == 'N/A':
        return 0
    if val == '10+':
        return 10
    try:
        int_val = int(val)
        return 10 if int_val >= 10 else int_val
    except ValueError:
        return 0
    
def analyze_stock(wb):
    sheetnames = wb.sheetnames
    
    for i in range(2, len(sheetnames)):
        prev_sheet = wb[sheetnames[i - 1]]
        curr_sheet = wb[sheetnames[i]]

        def sheet_to_dict(sheet):
            data = {}
            for row in sheet.iter_rows(min_row=2):
                name = str(row[1].value).strip() if row[1].value else None
                if not name:
                    continue
                values = [normalize_cell_value(cell.value) for cell in row[2:31]]
                data[name] = values
            return data
        
        prev_data = sheet_to_dict(prev_sheet)

        for row in curr_sheet.iter_rows(min_row=2):
            product_name = str(row[1].value).strip() if row[1].value else None
            if not product_name or product_name not in prev_data:
                continue

            prev_row = prev_data[product_name]
            curr_row = [normalize_cell_value(cell.value) for cell in row[2:31]]

            for j, (prev_val, curr_val) in enumerate(zip(prev_row, curr_row), start=2):
                cell = row[j]
                if(prev_val == 'N/A' and isinstance(curr_val, int)) or (curr_val == 'N/A' and isinstance(prev_val, int)):
                    cell.fill = blue_fill
                elif isinstance(prev_val, int) and isinstance(curr_val, int):
                    if curr_val > prev_val:
                        cell.fill = green_fill
                    elif curr_val < prev_val:
                        cell.fill = yellow_fill

# Format the sheet
def format_new_sheet(ws):
    category_positions = {}
    row_start = 2
    for category, products in [("Chassis", chassis),
                               ("AIO Liquid CPU Cooler", coolers),
                               ("External", external),
                               ("Power Supply", power_supplies)]:
        if not products:
            continue
        row_end = row_start + len(products) - 1
        ws.merge_cells(start_row=row_start, start_column=1, end_row = row_end, end_column=1)
        ws[f"A{row_start}"] = category
        category_positions[category] = row_start
        for i, name in enumerate(products.keys(), start=row_start):
            ws[f"B{i}"] = name
        row_start = row_end + 1

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=42):
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    return category_positions

def product_sums(ws):
    def get_merged_value(ws, cell_ref):
        cell = ws[cell_ref]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws[merged_range.coord.split(":")[0]].value
        return cell.value

    for row in range(2, ws.max_row + 1):
        ws[f"AQ{row}"] = f"=SUM(C{row}:AP{row})"
        ws[f"AS{row}"] = f"=COUNTIF(C{row}:AP{row}, 0)"

    category_blocks = {}
    current_category = None
    current_start = None

    for row in range(2, ws.max_row + 1):
        category_value = get_merged_value(ws, f"A{row}")

        if category_value != current_category:
            # Close the previous category block
            if current_category is not None:
                category_blocks[current_category] = (current_start, row)
            # Start a new category block
            current_category = category_value
            current_start = row

    # Close the last category block
    if current_category is not None:
        category_blocks[current_category] = (current_start, ws.max_row + 1)

    # 3. Write category totals in AS at the top row of each category
    for category, (start_row, end_row) in category_blocks.items():
        total_rows = []
        for row in range(start_row, end_row):
            if not get_merged_value(ws, f"A{row}"):  # Stop if blank
                break
            total_rows.append(f"AQ{row}")

        if total_rows:
            ws[f"AR{start_row}"] = f"=SUM({','.join(total_rows)})"

def prepare_chart_data(wb):
    # Ensure Charts sheet exists in position 3
    if len(wb.sheetnames) < 2:
        while len(wb.sheetnames) < 2:
            wb.create_sheet()
    if wb.sheetnames[1] != "Charts":
        charts_ws = wb.create_sheet("Charts")
        wb._sheets.insert(1, wb._sheets.pop())
    else:
        charts_ws = wb.worksheets[1]

    # Clear existing content
    for row in charts_ws.iter_rows():
        for cell in row:
            cell.value = None

    weekly_sheets = wb.sheetnames[2:]
    all_categories = {}
    model_data = {}
    category_totals = {}
    week_labels = []

    # Collect weekly data
    for sheetname in weekly_sheets:
        ws = wb[sheetname]
        week_labels.append(sheetname)

        last_category = None

        for row in range(2, ws.max_row + 1):
            category_cell = ws[f"A{row}"].value
            if category_cell and category_cell.strip():
                last_category = category_cell.strip()
                # Normalize similar category names
                if "cooler" in last_category.lower():
                    last_category = "Coolers"
                elif "chassis" in last_category.lower():
                    last_category = "Chassis"
                elif "power" in last_category.lower():
                    last_category = "Power Supplies"
            elif not last_category:
                continue  # No category yet

            category = last_category
            model = ws[f"B{row}"].value
            if not model or str(model).strip().upper().startswith("UPDATED"):
                continue

            # Keep order of models in category
            if category not in all_categories:
                all_categories[category] = []
            if model not in all_categories[category]:
                all_categories[category].append(model)

            # Sum columns D-AP
            total_value = 0
            for col in range(4, 42):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, (int, float)):
                    total_value += cell_val

            # Store model data
            model_data.setdefault(category, {}).setdefault(model, []).append(total_value)

            # Store category totals
            category_totals.setdefault(category, [0] * len(weekly_sheets))
            category_totals[category][len(week_labels) - 1] += total_value

    # Write product data per category
    current_row = 1
    for category, models in all_categories.items():
        charts_ws.cell(row=current_row, column=1).value = f"{category} Stock Trends Data"
        start_row = current_row + 1
        charts_ws.cell(row=start_row, column=1).value = "Week"

        for i, week in enumerate(week_labels):
            charts_ws.cell(row=start_row + i + 1, column=1).value = week

        for j, model in enumerate(models):
            charts_ws.cell(row=start_row, column=j + 2).value = model
            model_values = model_data[category].get(model, [])
            for i, value in enumerate(model_values):
                charts_ws.cell(row=start_row + i + 1, column=j + 2).value = value

        current_row = start_row + len(week_labels) + 3

    # Write category totals
    charts_ws.cell(row=current_row, column=1).value = "Category Totals Data"
    start_row = current_row + 1
    charts_ws.cell(row=start_row, column=1).value = "Week"
    # Preserve category order as first encountered
    ordered_categories = list(category_totals.keys())
    for i, cat in enumerate(ordered_categories):
        charts_ws.cell(row=start_row, column=i + 2).value = cat

    for w, week in enumerate(week_labels):
        charts_ws.cell(row=start_row + w + 1, column=1).value = week
        for i, cat in enumerate(ordered_categories):
            charts_ws.cell(row=start_row + w + 1, column=i + 2).value = category_totals[cat][w]

    current_row = start_row + len(week_labels) + 3

    # Write store totals
    charts_ws.cell(row=current_row, column=1).value = "Store Totals Data"
    start_row = current_row + 1
    charts_ws.cell(row=start_row, column=1).value = "Week"
    for i, store in enumerate(store_map):
        charts_ws.cell(row=start_row, column=i + 2).value = store
    for w, sheetname in enumerate(weekly_sheets):
        ws = wb[sheetname]
        charts_ws.cell(row=start_row + w + 1, column=1).value = sheetname
        for i, store in enumerate(store_map):
            store_total = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=4 + i).value
                if isinstance(val, (int, float)):
                    store_total += val
            charts_ws.cell(row=start_row + w + 1, column=i + 2).value = store_total

def run_stock_tracker(target_wb, sheet_name):
    # Setup Selenium driver
    options = Options()
    options.add_argument("--headless")
    if DEBUG_MODE:
        options.add_argument("--enable-logging")   
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    # Setup worksheet
    ws = target_wb.create_sheet(title=sheet_name)
    headers = ["Product Category", "Model", "Online"] + store_map + ["INDIVIDUAL TOTALS", "CATEGORY TOTALS", "OUT OF STOCK"]
    ws.append(headers)

    row_num = 2

    debug_lines = []

    # Start scanning for URLs
    for category, products in [("Chassis", chassis),
                               ("Cooler", coolers),
                               ("External", external),
                               ("Power Supply", power_supplies)]:

        for name, url in products.items():
            if not url and DEBUG_MODE:
                print(f"Skipping {name}: URL is missing")
                continue

            if url == 0:
                print(f"Error fetching {name}: No webpage for this product.")
                row = [category, name, "N/A"] + ["N/A"] * len(store_map)
                ws.append(row)
                row_num += 1
                continue

            product_result = {"url": url, "online": None, "stores": {}}

            header_line = f"\n=== Checking stock for: {name} ==="
            print(header_line)
            debug_lines.append(header_line)

            try:
                driver.get(url)
                time.sleep(2)

                online_status = "N/A"
                try:
                    online_element = driver.find_elements(By.XPATH, "//p[contains(@class, 'mt-1 text-dark f-16 fm-xs-SF-Pro-Display-Medium')]")
                    for element in online_element:
                        text = element.get_attribute("textContent").strip()
                        if "Available to Ship" in text:
                            online_status = "Yes"
                            break
                        else:
                            online_status = "No"
                            break
                except Exception as e:
                    print(f"[DEBUG] Could not detect online availability: {e}")
                    online_status = "N/A"

                product_result["online"] = online_status

                print(f"Online: {online_status}")
                debug_lines.append(f"Online: {online_status}")

                store_stock_map = {}
                for store in store_map:
                    try:
                        store_element = driver.find_elements(By.XPATH, f"//span[contains(text(), '{store}')]")
                        if store_element:
                            stock_element = store_element[0].find_element(By.XPATH, "following::span[contains(@class, 'shop-online-box')][1]")
                            raw_html = stock_element.get_attribute("outerHTML")
                            classes = stock_element.get_attribute("class")
                            raw_text = stock_element.get_attribute("textContent").strip()

                            if "bg-0000001" in classes:
                                stock_value = 0
                            elif "bg-E3E9F8" in classes:
                                # Try to parse number
                                if raw_text.isdigit():
                                    stock_value = int(raw_text)
                                else:
                                    stock_value = raw_text  # e.g. "10+"
                            else:
                                stock_value = raw_text

                            store_stock_map[store] = stock_value
                            
                            line = f"   → {store}: {raw_html} → {stock_value}"
                            print(line)
                            debug_lines.append(line)
                        else:
                            msg = f"   → {store}: NOT FOUND"
                            print(msg)
                            debug_lines.append(msg)
                            store_stock_map[store] = 0
                    except Exception:
                        msg = f"   → {store}: ERROR ({e})"
                        print(msg)
                        debug_lines.append(msg)
                        store_stock_map[store] = "N/A"
                
                print(f"{name}: {store_stock_map}")

                row = [category, name, online_status] + [store_stock_map.get(store, "N/A") for store in store_map]
                ws.append(row)
                row_num += 1
            
            except Exception as e:
                print(f"Error fetching {name}: {e}")
                row = [category, name] + ["N/A"] * len(store_map)
                ws.append(row)
                row_num += 1

    format_new_sheet(ws)

    product_sums(ws)

    driver.quit()

    # Save debug log to file
    with open("canada_computers_debug.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(debug_lines))

    print("[DEBUG] Saved raw scraping results to canada_computers_debug.json")

def terminate():
    sys.exit()

# Prompt user to add or remove products
def modify_products_window(use_original=False):
    filename = original_file if use_original else json_file

    win = tk.Toplevel()
    win.title("Modify Original Products" if use_original else "Modify Products")
    win.geometry("1200x700")
    win.protocol("WM_DELETE_WINDOW", terminate)

    with open(filename, "r") as f:
        data = json.load(f)
        local_chassis = data.get("chassis", {})
        local_coolers = data.get("coolers", {})
        local_external = data.get("external", {})
        local_power = data.get("power_supplies", {})

    def save_changes():
        with open(filename, "w") as f:
            json.dump({
                "chassis": local_chassis,
                "coolers": local_coolers,
                "external": local_external,
                "power_supplies": local_power
            }, f, indent=2)

    def refresh():
        if not win.winfo_exists():
            return
        
        all_items, idx = [], 1
        for category, products in [("Chassis", local_chassis),
                                   ("Coolers", local_coolers),
                                   ("External", local_external),
                                   ("Power Supplies", local_power)]:
            for name, url in products.items():
                all_items.append((idx, category, name, url))
                idx += 1

        names_text.config(state="normal")
        names_text.delete("1.0", "end")
        names_text.insert("1.0", "\n".join(f"{i}. [{cat}] {name}" for i, cat, name, _ in all_items))
        names_text.config(state="disabled")

        urls_text.config(state="normal")
        urls_text.delete("1.0", "end")
        urls_text.insert("1.0", "\n".join(f"{i}. {url}" for i, _, _, url in all_items))
        urls_text.config(state="disabled")

        return all_items

    # Display model names
    tk.Label(win, text="Original Product Names" if use_original else "Product Names", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
    names_text = tk.Text(win, width=40, height=25, wrap="word")
    names_text.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
    names_text.config(state="disabled")

    # Display URLs
    tk.Label(win, text="Original Product URLs" if use_original else "Product URLs", font=("Arial", 12, "bold")).grid(row=0, column=1, padx=10, pady=5, sticky="w")
    urls_text = tk.Text(win, width=120, height=25, wrap="word")
    urls_text.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")
    urls_text.config(state="disabled")

    # Add and Remove buttons
    def add_product():
        add_win = tk.Toplevel(win)
        add_win.title("Add New Product")
        add_win.geometry("600x200")
        add_win.protocol("WM_DELETE_WINDOW", terminate)

        tk.Label(add_win, text="Model Name:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        model_entry = tk.Entry(add_win, width=50)
        model_entry.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(add_win, text="Product URL:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        url_entry = tk.Entry(add_win, width=50)
        url_entry.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(add_win, text="Category:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        category_var = tk.StringVar(value="Power Supply")
        tk.OptionMenu(add_win, category_var, "Chassis", "Cooler", "External", "Power Supply").grid(row=2, column=1, padx=10, pady=10, sticky="w")

        def confirm_add():
            new_model, new_url, category = model_entry.get().strip(), url_entry.get().strip(), category_var.get()

            if not new_model or not new_url:
                messagebox.showerror("Error", "Both Model Name and Product URL are required.")
                return
            if category == "Power Supply":
                local_power[new_model] = new_url
            elif category == "Cooler":
                local_coolers[new_model] = new_url
            elif category == "Chassis":
                local_chassis[new_model] = new_url
            elif category == "External":
                local_external[new_model] = new_url
            save_changes()
            messagebox.showinfo("Product Added", f"{new_model} added to {category}.")
            add_win.destroy()
            refresh()

        tk.Button(add_win, text="Add", command=confirm_add, width=15).grid(row=2, column=0, pady=20)
        tk.Button(add_win, text="Cancel", command=add_win.destroy, width=15).grid(row=2, column=1, pady=20)

    def remove_product():
        all_items = refresh()

        remove_model = simpledialog.askstring("Remove Product", "Enter the exact model name or model number to remove:")
        if not remove_model:
            return
        
        if remove_model.isdigit():
            remove_idx = int(remove_model)
            for idx, category, name, _ in all_items:
                if idx == remove_idx:
                    target_map = {
                        "Chassis": local_chassis,
                        "Coolers": local_coolers,
                        "Miscellaneous": local_external,
                        "Power Supplies": local_power
                    }[category]
                    del target_map[name]
                    save_changes()
                    messagebox.showinfo("Product Removed", f"{name} removed.")
                    refresh()
                    return
        else:
            for product_map in [local_chassis, local_coolers, local_external, local_power]:
                if remove_model in product_map:
                    del product_map[remove_model]
                    save_changes()
                    messagebox.showinfo("Product Removed", f"{remove_model} removed.")
                    refresh()
                    return

        messagebox.showinfo("Not Found", f"{remove_model} not found.")

    def reset_to_original():
        if not os.path.exists("CC_original_products.json"):
            messagebox.showerror("Error", "CC_original_products.json not found.")
            return
        with open(original_file, "r") as f:
            data = json.load(f)
            chassis.clear(); chassis.update(data.get("chassis", {}))
            coolers.clear(); coolers.update(data.get("coolers", {}))
            external.clear(); external.update(data.get("external", {}))
            power_supplies.clear(); power_supplies.update(data.get("power_supplies", {}))
            # Also update the local GUI references
            local_chassis.clear(); local_chassis.update(chassis)
            local_coolers.clear(); local_coolers.update(coolers)
            local_external.clear(); local_external.update(external)
            local_power.clear(); local_power.update(power_supplies)

        save_products(json_file)
        refresh()
        messagebox.showinfo("Reset Complete", "Product list reset to original.")


    def done():
        win.quit()
        win.destroy()

    tk.Button(win, text="Add Product", command=add_product, width=20).grid(row=2, column=0, pady=10)
    tk.Button(win, text="Remove Product", command=remove_product, width=20).grid(row=2, column=1, pady=10)
    tk.Button(win, text="Done", command=done, width=20).grid(row=3, column=0, pady=10)
    if not use_original:  
        tk.Button(win, text="Reset to Original", command=reset_to_original, width=20).grid(row=3, column=1, pady=10)
        tk.Button(win, text="Modify Original", command=lambda: modify_products_window(use_original=True), width=20).grid(row=4, column=1, pady=10)

    refresh()
    win.mainloop()

def main():
    load_products()
    final_message = "Error"

    # Prompt user for file save location
    root = Tk()
    root.withdraw()  # Hide the main window
    root.protocol("WM_DELETE_WINDOW", terminate)

    # Ask the user for the week number (default is 1)
    week_number = simpledialog.askstring("Week Number", "Enter the week number (e.g., 9, 10, 11, etc):")
    if not week_number:
        terminate()

    if messagebox.askyesno("Modify Products", "Do you want to view, add, or remove products?"):
        modify_products_window()

    use_existing = messagebox.askyesno("Stock Tracker", "Would you like to edit an existing Excel file?")

    file_path = None
    # If the user owns the stock tracker file
    if use_existing:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            messagebox.showinfo("Cancelled", "No file selected. Exiting.")
            time.sleep(2)
            terminate()
    else:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="Save new stock report as")
        if not file_path:
            messagebox.showinfo("Cancelled", "No save location selected.")
            time.sleep(2)
            terminate()

    global STOCK_TRACKER_START
    STOCK_TRACKER_START = time.time()

    analyze_only = messagebox.askyesno("ANALYZE ONLY", "Would you like to ONLY run the highlighting/labeling program?")

    if use_existing:
        # Run the stock tracker and coloring
        wb = load_workbook(file_path)
        sheet_name = f"WK{week_number}"
        process_add_products_sheet(wb)
        if not analyze_only:
            run_stock_tracker(wb, sheet_name)
        prepare_chart_data(wb)
    else:
        # If the user opts to create an independent sheet with this week's stock
        wb = Workbook()
        ws = wb.active
        sheet_name = f"WK{week_number}"
        ws.title = sheet_name
        run_stock_tracker(wb, sheet_name)

    analyze_stock(wb)
    wb.save(file_path)
    elapsed = time.time() - STOCK_TRACKER_START
    processed_time = f"{int(elapsed // 3600)} Hours, {int((elapsed % 3600) // 60)} Minutes, {int((elapsed % 3600) % 60)} Seconds"
    final_message = f"Stock tracking completed in {processed_time}"
    print(final_message)
    os.startfile(file_path)

    time.sleep(2)
    terminate()

if __name__ == "__main__":
    main()